"""
audit.py — Excel data quality analysis for Excel MCP v2.
Same checks as v1. Now importable standalone without server.py dependency.

Two public functions:

  audit_workbook(ai_path, graph) → dict of all issues found
  apply_fixes(ai_path, fixes)    → applies LLM-proposed fixes to ai_workbook

Issue types:
  TYPE_VIOLATION      — cell value doesn't match column's declared type
  FORMAT_VIOLATION    — value in wrong format for declared number_format
                        (e.g. 18 in a 0.00% column instead of 0.18)
  RANGE_VIOLATION     — value outside valid range (pct > 1, negative budget, etc.)
  FORMULA_OVERRIDE    — formula column cell has hardcoded value (silent corruption)
  MISSING_REQUIRED    — null where column has < 5% nulls overall
  DUPLICATE_ROW       — exact duplicate of another row
  LOGIC_VIOLATION     — sum of parts ≠ total (detectable cross-column inconsistency)
  CONSISTENCY         — value doesn't match pattern of rest of column
"""

import re, json, io, base64, textwrap
from pathlib import Path
from collections import defaultdict, Counter
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

# ─────────────────────────────────────────────────────────────────────────────
# ISSUE SEVERITY
# ─────────────────────────────────────────────────────────────────────────────

SEVERITY = {
    "FORMULA_OVERRIDE":  "CRITICAL",   # silent data corruption
    "LOGIC_VIOLATION":   "HIGH",
    "TYPE_VIOLATION":    "HIGH",
    "FORMAT_VIOLATION":  "MEDIUM",
    "RANGE_VIOLATION":   "MEDIUM",
    "MISSING_REQUIRED":  "MEDIUM",
    "DUPLICATE_ROW":     "LOW",
    "CONSISTENCY":       "LOW",
}

SEVERITY_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _is_numeric_str(v) -> bool:
    try: float(str(v).replace(",","").replace("₹","").replace("%","").strip()); return True
    except: return False

def _looks_like_date(v) -> bool:
    if hasattr(v, "year"): return True
    s = str(v).strip()
    patterns = [r"\d{2}[/-]\d{2}[/-]\d{2,4}", r"\d{4}[/-]\d{2}[/-]\d{2}",
                r"\d{1,2}\s+\w{3}\s+\d{4}"]
    return any(re.match(p, s) for p in patterns)

def _infer_expected(col_info: dict, col_type: str) -> str:
    fmt = col_info.get("format_meaning", "general")
    if fmt == "percentage": return "decimal between 0–1 (e.g. 0.75 for 75%)"
    if fmt in ("currency_inr","currency_usd"): return "positive integer/float"
    if fmt == "date": return "date value or serial number"
    if col_type == "numeric": return "numeric value"
    if col_type == "text": return "text string"
    return "non-empty value"


# ─────────────────────────────────────────────────────────────────────────────
# FORMULA PATTERN EXTRACTION from graph
# ─────────────────────────────────────────────────────────────────────────────

def _formula_columns(sheet_graph: dict) -> dict:
    """
    Returns {col_letter: example_formula} for columns whose majority cells are formulas.
    E.g. {"J": "=SUM(F2:I2)", "K": "=IF(E2>0,J2/E2,0)"}
    """
    result = {}
    patterns = sheet_graph.get("formulas", {}).get("patterns", [])
    for p in patterns:
        # Extract column letter from example_cell (e.g. "J2" → "J")
        m = re.match(r"([A-Z]+)\d+", p["example_cell"])
        if m:
            col = m.group(1)
            # Only mark as formula-column if it appears ≥3 times (not a one-off)
            if p["count"] >= 3:
                result[col] = p["formula"]
    return result


# ─────────────────────────────────────────────────────────────────────────────
# CORE AUDIT
# ─────────────────────────────────────────────────────────────────────────────

def audit_workbook(ai_path: str, graph: dict) -> dict:
    """
    Full structural + data quality audit.
    Returns:
    {
      "summary": {total, by_severity, by_sheet, by_type},
      "issues": [ {sheet, row, col_letter, column_name, issue_type, severity,
                   found_value, expected, suggestion}, ... ],
      "clean_columns": {sheet: [col_names with no issues]},
      "stats": {sheet: {row_count, null_count, duplicate_count, formula_override_count}}
    }
    """
    issues = []

    p = Path(ai_path)

    # Open formula workbook (to detect hardcoded overrides in formula columns)
    wb_f = openpyxl.load_workbook(str(p), data_only=False, read_only=True)
    # Open values workbook (to check actual computed/entered values)
    wb_v = openpyxl.load_workbook(str(p), data_only=True, read_only=True)

    sheet_stats = {}

    for sname, s_graph in graph.get("sheets", {}).items():
        if sname not in wb_f.sheetnames:
            continue

        ws_f = wb_f[sname]
        ws_v = wb_v[sname]

        headers   = [str(h) if h is not None else f"Col_{i+1}"
                     for i, h in enumerate(s_graph.get("headers", []))]
        col_types = s_graph.get("column_types", {})
        col_fmts  = s_graph.get("column_formats", {})
        max_row   = s_graph.get("max_row", 0)
        formula_cols = _formula_columns(s_graph)  # {col_letter: formula_pattern}

        # ── Read all data values ───────────────────────────────────────────────
        # For large sheets: we still need to scan every cell.
        # We use read_only=True + row iteration (no random access) — memory-safe.

        col_null_counts   = defaultdict(int)
        col_total_counts  = defaultdict(int)
        col_sample_values = defaultdict(list)  # for consistency checking
        row_hashes        = []                  # for duplicate detection
        formula_override_cells = []

        data_rows = list(ws_v.iter_rows(min_row=2, values_only=True))
        formula_rows = list(ws_f.iter_rows(min_row=2, values_only=False))

        n_data_rows = len(data_rows)

        for ri, (v_row, f_row) in enumerate(zip(data_rows, formula_rows), start=2):
            row_vals = []
            for ci, (v_cell_val, f_cell) in enumerate(
                    zip(v_row, f_row), start=1):

                col_letter = get_column_letter(ci)
                col_name   = headers[ci-1] if ci <= len(headers) else col_letter
                col_type   = col_types.get(col_name, "unknown")
                fmt_info   = col_fmts.get(col_name, {})
                fmt_meaning= fmt_info.get("format_meaning", "general")
                hidden     = fmt_info.get("hidden", False)

                row_vals.append(str(v_cell_val))
                col_total_counts[col_name] += 1

                # Skip hidden columns from most checks (they're internal)
                if hidden:
                    continue

                # ── 1. FORMULA OVERRIDE check ──────────────────────────────────
                if col_letter in formula_cols:
                    f_val = f_cell.value
                    if f_val is not None and not (isinstance(f_val, str) and f_val.startswith("=")):
                        # Cell has a hardcoded value where formula expected
                        issues.append({
                            "sheet": sname, "row": ri,
                            "col_letter": col_letter, "column_name": col_name,
                            "issue_type": "FORMULA_OVERRIDE",
                            "severity": "CRITICAL",
                            "found_value": repr(f_val),
                            "expected": f"formula like `{formula_cols[col_letter]}`",
                            "suggestion": f"Replace with formula: `{formula_cols[col_letter].replace('2', str(ri))}`",
                        })
                        formula_override_cells.append((ri, col_letter))
                        continue  # don't double-report

                # ── 2. MISSING REQUIRED check ──────────────────────────────────
                if v_cell_val is None or (isinstance(v_cell_val, str) and v_cell_val.strip() == ""):
                    col_null_counts[col_name] += 1
                    # Will decide if MISSING_REQUIRED after full pass
                    continue

                val = v_cell_val

                # Collect samples for consistency
                if len(col_sample_values[col_name]) < 200:
                    col_sample_values[col_name].append(val)

                # ── 3. TYPE VIOLATION check ────────────────────────────────────
                if col_type == "numeric" and fmt_meaning not in ("date", "time"):
                    if isinstance(val, str):
                        s = val.strip()
                        if s and not _is_numeric_str(s):
                            issues.append({
                                "sheet": sname, "row": ri,
                                "col_letter": col_letter, "column_name": col_name,
                                "issue_type": "TYPE_VIOLATION",
                                "severity": SEVERITY["TYPE_VIOLATION"],
                                "found_value": repr(val),
                                "expected": _infer_expected(fmt_info, col_type),
                                "suggestion": f"Convert to number or replace with 0",
                            })

                elif col_type == "text":
                    if isinstance(val, (int, float)) and col_name.lower() not in (
                            "sl_no","id","code","number","no"):
                        # Numeric in text column — might be OK (codes), flag as low severity
                        pass  # too noisy, skip

                # ── 4. FORMAT VIOLATION check ──────────────────────────────────
                if fmt_meaning == "percentage":
                    if isinstance(val, (int, float)):
                        if val > 1.5:  # clearly entered as 75 instead of 0.75
                            issues.append({
                                "sheet": sname, "row": ri,
                                "col_letter": col_letter, "column_name": col_name,
                                "issue_type": "FORMAT_VIOLATION",
                                "severity": SEVERITY["FORMAT_VIOLATION"],
                                "found_value": repr(val),
                                "expected": "decimal 0–1 (column is formatted as %)",
                                "suggestion": f"Divide by 100 → {val/100:.4f}",
                            })

                elif fmt_meaning in ("currency_inr", "currency_usd", "number_thousands"):
                    if isinstance(val, str) and _is_numeric_str(val):
                        # Stored as text but should be number
                        clean = val.replace(",","").replace("₹","").replace("$","").strip()
                        issues.append({
                            "sheet": sname, "row": ri,
                            "col_letter": col_letter, "column_name": col_name,
                            "issue_type": "FORMAT_VIOLATION",
                            "severity": SEVERITY["FORMAT_VIOLATION"],
                            "found_value": repr(val),
                            "expected": "numeric value (not text string)",
                            "suggestion": f"Convert to number: {clean}",
                        })

                # ── 5. RANGE VIOLATION check ───────────────────────────────────
                if fmt_meaning == "percentage" and isinstance(val, (int, float)):
                    if val < 0 or val > 1.5:
                        issues.append({
                            "sheet": sname, "row": ri,
                            "col_letter": col_letter, "column_name": col_name,
                            "issue_type": "RANGE_VIOLATION",
                            "severity": SEVERITY["RANGE_VIOLATION"],
                            "found_value": repr(val),
                            "expected": "value between 0 and 1",
                            "suggestion": f"Check: should this be {abs(val)/100:.4f}?" if abs(val) > 1 else "Remove negative value",
                        })

                elif fmt_meaning in ("currency_inr", "currency_usd") and isinstance(val, (int, float)):
                    if val < 0:
                        issues.append({
                            "sheet": sname, "row": ri,
                            "col_letter": col_letter, "column_name": col_name,
                            "issue_type": "RANGE_VIOLATION",
                            "severity": SEVERITY["RANGE_VIOLATION"],
                            "found_value": repr(val),
                            "expected": "non-negative currency amount",
                            "suggestion": "Check: may be a data entry error. Use absolute value or investigate.",
                        })

            # Row hash for duplicate detection
            row_hashes.append(tuple(row_vals))

        # ── 6. MISSING REQUIRED (post-pass) ───────────────────────────────────
        for col_name, null_count in col_null_counts.items():
            total = col_total_counts.get(col_name, 1)
            null_pct = null_count / total
            if null_pct < 0.05:
                pass  # column almost always full — individual nulls already noted
            elif null_pct < 0.5 and null_count >= 3:
                # Many nulls in a moderately-filled column — report as column-level
                issues.append({
                    "sheet": sname, "row": None,
                    "col_letter": None, "column_name": col_name,
                    "issue_type": "MISSING_REQUIRED",
                    "severity": SEVERITY["MISSING_REQUIRED"],
                    "found_value": f"{null_count} nulls ({null_pct:.0%} of {total} rows)",
                    "expected": "populated values",
                    "suggestion": f"Review column `{col_name}`: {null_count} cells are empty",
                })

        # Report individual missing values in columns that are otherwise full
        for col_name, null_count in col_null_counts.items():
            total = col_total_counts.get(col_name, 1)
            null_pct = null_count / total
            # Only flag individual cells if column is ≥ 90% full
            if null_pct < 0.1 and null_count <= 10:
                # Find the actual rows (re-scan — only for small null counts)
                for ri, v_row in enumerate(data_rows, start=2):
                    ci = headers.index(col_name) + 1 if col_name in headers else None
                    if ci and ci <= len(v_row):
                        val = v_row[ci-1]
                        if val is None or (isinstance(val, str) and val.strip() == ""):
                            issues.append({
                                "sheet": sname, "row": ri,
                                "col_letter": get_column_letter(ci),
                                "column_name": col_name,
                                "issue_type": "MISSING_REQUIRED",
                                "severity": SEVERITY["MISSING_REQUIRED"],
                                "found_value": "null/empty",
                                "expected": _infer_expected(
                                    col_fmts.get(col_name, {}),
                                    col_types.get(col_name, "unknown")),
                                "suggestion": "Fill in missing value",
                            })

        # ── 7. DUPLICATE ROWS ──────────────────────────────────────────────────
        hash_counts = Counter(row_hashes)
        seen = set()
        for ri, h in enumerate(row_hashes, start=2):
            if hash_counts[h] > 1 and h not in seen:
                seen.add(h)
                issues.append({
                    "sheet": sname, "row": ri,
                    "col_letter": None, "column_name": "(entire row)",
                    "issue_type": "DUPLICATE_ROW",
                    "severity": SEVERITY["DUPLICATE_ROW"],
                    "found_value": f"Row {ri} duplicated {hash_counts[h]} times",
                    "expected": "unique rows",
                    "suggestion": f"Delete duplicate rows. First occurrence at row {ri}.",
                })

        # ── 8. LOGIC VIOLATION: sum-of-parts check ────────────────────────────
        # If we can detect a SUM formula pattern (J = SUM(F:I)), verify it
        for p in s_graph.get("formulas", {}).get("patterns", []):
            formula = p.get("formula", "")
            m = re.match(r"=SUM\(([A-Z]+)\d+:([A-Z]+)\d+\)", formula)
            if m:
                result_col   = re.match(r"([A-Z]+)", p["example_cell"]).group(1)
                start_col    = m.group(1)
                end_col      = m.group(2)
                start_ci     = column_index_from_string(start_col)
                end_ci       = column_index_from_string(end_col)
                result_ci    = column_index_from_string(result_col)
                result_name  = headers[result_ci-1] if result_ci <= len(headers) else result_col

                for ri, v_row in enumerate(data_rows, start=2):
                    if result_ci > len(v_row): continue
                    total_val = v_row[result_ci - 1]
                    if total_val is None: continue
                    if not isinstance(total_val, (int, float)): continue
                    parts = [v_row[ci-1] for ci in range(start_ci, end_ci+1) if ci <= len(v_row)]
                    parts_numeric = [x for x in parts if isinstance(x, (int, float))]
                    if len(parts_numeric) == (end_ci - start_ci + 1):
                        expected_total = sum(parts_numeric)
                        if abs(expected_total - total_val) > 1:  # 1 unit tolerance for rounding
                            issues.append({
                                "sheet": sname, "row": ri,
                                "col_letter": result_col,
                                "column_name": result_name,
                                "issue_type": "LOGIC_VIOLATION",
                                "severity": SEVERITY["LOGIC_VIOLATION"],
                                "found_value": f"{total_val:,.2f}",
                                "expected": f"SUM({start_col}:{end_col}) = {expected_total:,.2f}",
                                "suggestion": f"Row {ri}: {result_name} should be {expected_total:,.2f}, found {total_val:,.2f}. Re-enter formula.",
                            })
                break  # only check first SUM pattern per sheet

        sheet_stats[sname] = {
            "rows": n_data_rows,
            "issues": sum(1 for iss in issues if iss["sheet"] == sname),
            "null_cells": sum(col_null_counts.values()),
            "duplicate_rows": sum(1 for iss in issues if iss["sheet"] == sname and iss["issue_type"] == "DUPLICATE_ROW"),
            "formula_overrides": len(formula_override_cells),
        }

    wb_f.close()
    wb_v.close()

    # ── Build summary ──────────────────────────────────────────────────────────
    by_severity = Counter(iss["severity"] for iss in issues)
    by_type     = Counter(iss["issue_type"] for iss in issues)
    by_sheet    = Counter(iss["sheet"] for iss in issues)

    # Columns with zero issues per sheet
    clean_columns = {}
    for sname, s_graph in graph.get("sheets", {}).items():
        headers = [str(h) for h in s_graph.get("headers", []) if h is not None]
        dirty = {iss["column_name"] for iss in issues if iss["sheet"] == sname}
        clean_columns[sname] = [h for h in headers if h not in dirty]

    # Sort issues: CRITICAL first, then by sheet/row
    issues.sort(key=lambda x: (
        SEVERITY_ORDER.get(x["severity"], 9),
        x["sheet"],
        x["row"] or 0
    ))

    return {
        "summary": {
            "total_issues": len(issues),
            "by_severity": dict(by_severity),
            "by_type": dict(by_type),
            "by_sheet": dict(by_sheet),
        },
        "issues": issues,
        "clean_columns": clean_columns,
        "sheet_stats": sheet_stats,
    }


# ─────────────────────────────────────────────────────────────────────────────
# HTML AUDIT REPORT
# ─────────────────────────────────────────────────────────────────────────────

def render_audit_html(audit_result: dict, workbook_name: str, out_path: Path) -> str:
    issues = audit_result["issues"]
    summary = audit_result["summary"]

    # Severity badge colours
    sev_colours = {
        "CRITICAL": ("#c0392b","#fdf0ef"),
        "HIGH":     ("#e67e22","#fef5ec"),
        "MEDIUM":   ("#f39c12","#fefce8"),
        "LOW":      ("#27ae60","#eafaf1"),
    }

    # Summary pills
    pills = ""
    for sev in ["CRITICAL","HIGH","MEDIUM","LOW"]:
        count = summary["by_severity"].get(sev, 0)
        if count:
            fg, bg = sev_colours[sev]
            pills += f'<span class="pill" style="background:{bg};color:{fg};border:1px solid {fg}">{sev}: {count}</span> '

    # Issue rows
    sev_cards_html = ""
    for sev in ["CRITICAL","HIGH","MEDIUM","LOW"]:
        fg, bg = sev_colours.get(sev, ("#333","#fff"))
        cnt = summary["by_severity"].get(sev, 0)
        sev_cards_html += f'<div class="card"><h3>{sev}</h3><div class="big" style="color:{fg}">{cnt}</div></div>'
        rows_html = ""
    for iss in issues:
        sev = iss["severity"]
        fg, bg = sev_colours.get(sev, ("#333","#fff"))
        sev_badge = f'<span class="badge" style="background:{bg};color:{fg};border:1px solid {fg}">{sev}</span>'
        cell_ref = f"{iss.get('col_letter','')}{iss.get('row','')}" if iss.get("row") else "column-level"
        rows_html += f"""<tr>
  <td>{iss['sheet']}</td>
  <td style="font-family:monospace">{cell_ref}</td>
  <td>{iss['column_name']}</td>
  <td><code>{iss['issue_type']}</code></td>
  <td>{sev_badge}</td>
  <td style="color:#c0392b;font-family:monospace">{str(iss['found_value'])[:60]}</td>
  <td style="color:#27ae60;font-size:0.82em">{str(iss['expected'])[:60]}</td>
  <td style="font-size:0.82em">{str(iss['suggestion'])[:80]}</td>
</tr>\n"""

    # Issue type breakdown
    type_breakdown = "".join(
        f"<li><code>{t}</code>: {c}</li>"
        for t, c in sorted(summary["by_type"].items(), key=lambda x: -x[1])
    )

    html = textwrap.dedent(f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Audit: {workbook_name}</title>
<style>
  body {{ font-family: 'Segoe UI', sans-serif; margin: 24px; background: #f5f7fa; color: #1a1a2e; }}
  h1 {{ color: #1F3864; font-size: 1.5em; margin-bottom: 4px; }}
  .meta {{ color:#666; font-size:0.85em; margin-bottom: 16px; }}
  .pills {{ margin-bottom: 16px; }}
  .pill {{ display:inline-block; padding:4px 10px; border-radius:12px;
           font-size:0.82em; font-weight:bold; margin-right:6px; }}
  .badge {{ display:inline-block; padding:2px 8px; border-radius:8px;
            font-size:0.8em; font-weight:bold; }}
  .summary-grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(200px,1fr));
    gap:12px; margin-bottom:20px; }}
  .card {{ background:white; border-radius:8px; padding:14px; box-shadow:0 1px 4px rgba(0,0,0,0.07); }}
  .card h3 {{ margin:0 0 8px; font-size:0.9em; color:#666; }}
  .card .big {{ font-size:2em; font-weight:bold; color:#1F3864; }}
  .search-box {{ margin-bottom:12px; }}
  input#searchInput {{ padding:8px 12px; width:360px; border:1px solid #ccc;
    border-radius:6px; font-size:0.9em; }}
  select.filter-sel {{ padding:8px; border:1px solid #ccc; border-radius:6px;
    font-size:0.88em; margin-left:8px; }}
  table {{ border-collapse:collapse; width:100%; background:white;
    box-shadow:0 1px 4px rgba(0,0,0,0.08); border-radius:8px; overflow:hidden; }}
  th {{ background:#1F3864; color:white; padding:10px 10px; text-align:left;
    cursor:pointer; font-size:0.82em; white-space:nowrap; }}
  th:hover {{ background:#2E5090; }}
  td {{ padding:7px 10px; font-size:0.81em; border-bottom:1px solid #eee;
    vertical-align:top; }}
  tr:hover td {{ background:#fff8e1; }}
  code {{ background:#f0f0f0; padding:1px 4px; border-radius:3px; font-size:0.9em; }}
  .type-list {{ columns:2; list-style:none; padding:0; font-size:0.85em; }}
  .count-badge {{ background:#1F3864;color:white;border-radius:10px;
    padding:2px 8px; font-size:0.78em; margin-left:6px; }}
</style>
</head>
<body>
<h1>📊 Data Quality Audit — {workbook_name}</h1>
<p class="meta">Total issues found: <strong>{summary['total_issues']}</strong></p>
<div class="pills">{pills}</div>

<div class="summary-grid">
  <div class="card"><h3>Total Issues</h3><div class="big">{summary['total_issues']}</div></div>
  {sev_cards_html}
</div>

<details open style="margin-bottom:16px">
  <summary style="cursor:pointer;font-weight:bold;color:#1F3864">Issue Type Breakdown</summary>
  <ul class="type-list" style="margin-top:8px">{type_breakdown}</ul>
</details>

<div class="search-box">
  <input type="text" id="searchInput" onkeyup="filterTable()" placeholder="Search issues...">
  <select class="filter-sel" id="sevFilter" onchange="filterTable()">
    <option value="">All severities</option>
    <option>CRITICAL</option><option>HIGH</option>
    <option>MEDIUM</option><option>LOW</option>
  </select>
  <select class="filter-sel" id="typeFilter" onchange="filterTable()">
    <option value="">All types</option>
    {"".join(f'<option>{t}</option>' for t in sorted(summary["by_type"].keys()))}
  </select>
</div>

<table id="issueTable">
  <thead><tr>
    <th>Sheet</th><th>Cell</th><th>Column</th><th>Issue Type</th>
    <th>Severity</th><th>Found Value</th><th>Expected</th><th>Suggestion</th>
  </tr></thead>
  <tbody id="tableBody">{rows_html}</tbody>
</table>
<p id="rowCount" style="color:#666;font-size:0.85em;margin-top:8px">{len(issues)} issues shown</p>

<script>
function filterTable() {{
  var inp = document.getElementById("searchInput").value.toLowerCase();
  var sev = document.getElementById("sevFilter").value;
  var typ = document.getElementById("typeFilter").value;
  var rows = document.getElementById("tableBody").getElementsByTagName("tr");
  var vis = 0;
  for (var r of rows) {{
    var txt = r.textContent;
    var show = (!inp || txt.toLowerCase().includes(inp)) &&
               (!sev  || txt.includes(sev)) &&
               (!typ  || txt.includes(typ));
    r.style.display = show ? "" : "none";
    if (show) vis++;
  }}
  document.getElementById("rowCount").textContent = vis + " issues shown";
}}
</script>
</body></html>""")

    out_path.write_text(html, encoding="utf-8")
    return str(out_path)


# ─────────────────────────────────────────────────────────────────────────────
# AUDIT SUMMARY for LLM (compact, structured for LLM reasoning)
# ─────────────────────────────────────────────────────────────────────────────

def render_audit_md(audit_result: dict, html_path: str, json_path: str) -> str:
    summary = audit_result["summary"]
    issues  = audit_result["issues"]
    stats   = audit_result.get("sheet_stats", {})

    L = []; a = L.append

    a("## Data Quality Audit Report\n")

    # Severity summary
    a("### Issue Counts by Severity")
    for sev in ["CRITICAL","HIGH","MEDIUM","LOW"]:
        n = summary["by_severity"].get(sev, 0)
        if n: a(f"- **{sev}**: {n}")
    a(f"\n**Total: {summary['total_issues']} issues**\n")

    # By type
    a("### By Issue Type")
    for t, c in sorted(summary["by_type"].items(), key=lambda x: -x[1]):
        a(f"- `{t}`: {c}")
    a("")

    # Sheet stats
    if stats:
        a("### Per-Sheet Stats")
        for sname, s in stats.items():
            a(f"- **{sname}**: {s['rows']} rows | {s['issues']} issues "
              f"| {s['null_cells']} null cells "
              f"| {s['formula_overrides']} formula overrides")
        a("")

    # Top issues (max 30 for LLM context)
    a("### Issues (sorted by severity)")
    a("*Full list in HTML report. Showing top 30 for LLM context.*\n")
    a("| Sheet | Cell | Column | Type | Severity | Found | Suggestion |")
    a("|---|---|---|---|---|---|---|")
    for iss in issues[:30]:
        cell = f"{iss.get('col_letter','')}{iss.get('row','')}" if iss.get("row") else "col-level"
        found = str(iss["found_value"])[:40].replace("|","\\|")
        sugg  = str(iss["suggestion"])[:60].replace("|","\\|")
        a(f"| {iss['sheet']} | {cell} | {iss['column_name']} | `{iss['issue_type']}` | "
          f"**{iss['severity']}** | {found} | {sugg} |")

    if len(issues) > 30:
        a(f"\n*... and {len(issues)-30} more issues. See HTML report.*")

    a(f"\n📄 **Full audit report (sortable):** `{html_path}`")
    a(f"📋 **Machine-readable JSON:** `{json_path}`")
    a(f"\n**Next steps:**")
    a("1. Review CRITICAL issues first (formula overrides)")
    a("2. Ask me to generate fixes: *'fix all FORMAT_VIOLATION issues in Expenditure sheet'*")
    a("3. Or: *'what are the correct values for the LOGIC_VIOLATION rows?'*")
    a("4. Then: `excel_apply_fixes(ai_path, fixes_json_path)` to apply")

    return "\n".join(L)


# ─────────────────────────────────────────────────────────────────────────────
# APPLY FIXES
# ─────────────────────────────────────────────────────────────────────────────

def apply_fixes(ai_path: str, fixes: list) -> dict:
    """
    Apply a list of fix dicts to the ai_workbook.

    Each fix:
    {
      "sheet": "Expenditure",
      "row": 5,
      "col_letter": "K",
      "new_value": 0.75,         # for value fix
      "new_formula": "=J5/E5",   # OR formula fix (takes priority)
    }

    Returns:
    {
      "applied": [...],
      "skipped": [...],
      "diff_summary": str,
    }
    """
    p = Path(ai_path)
    wb = openpyxl.load_workbook(str(p))

    applied = []
    skipped = []

    for fix in fixes:
        sheet = fix.get("sheet")
        row   = fix.get("row")
        col   = fix.get("col_letter")

        if not (sheet and row and col):
            skipped.append({**fix, "reason": "missing sheet/row/col"})
            continue
        if sheet not in wb.sheetnames:
            skipped.append({**fix, "reason": f"sheet '{sheet}' not found"})
            continue

        ws = wb[sheet]
        cell = ws[f"{col}{row}"]
        old_val = cell.value

        if "new_formula" in fix and fix["new_formula"]:
            cell.value = fix["new_formula"]
            applied.append({**fix, "old_value": repr(old_val), "action": "formula restored"})
        elif "new_value" in fix:
            cell.value = fix["new_value"]
            applied.append({**fix, "old_value": repr(old_val), "action": "value updated"})
        else:
            skipped.append({**fix, "reason": "no new_value or new_formula provided"})

    wb.save(str(p))

    diff_lines = [f"Applied {len(applied)} fixes, skipped {len(skipped)}:"]
    for a_fix in applied[:20]:
        diff_lines.append(
            f"  ✅ {a_fix['sheet']}!{a_fix.get('col_letter','')}{a_fix.get('row','')}: "
            f"{a_fix['old_value']} → {a_fix.get('new_formula') or repr(a_fix.get('new_value'))}"
        )
    if len(applied) > 20:
        diff_lines.append(f"  ... and {len(applied)-20} more")
    for s_fix in skipped:
        diff_lines.append(f"  ⚠️ Skipped: {s_fix.get('reason','?')}")

    return {
        "applied": applied,
        "skipped": skipped,
        "diff_summary": "\n".join(diff_lines),
    }
