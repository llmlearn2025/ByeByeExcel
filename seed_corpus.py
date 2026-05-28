"""
seed_corpus.py — Initial knowledge base entries for ByeByeExcel.

Run once to populate excel_knowledge.db with working examples.
Each entry is a real, tested pattern from authoritative sources.

Usage:
    python seed_corpus.py
    python seed_corpus.py --db /path/to/custom.db
    python seed_corpus.py --stats   # show what's in the DB

After running, the MCP can immediately answer LLM queries like:
  "how do I calculate a running total?"
  "group by month and show totals"
  "find rows where value is above average"
  "rolling 3-month average"
  "how to handle percentage columns correctly"
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))
from knowledge import init_knowledge_db, bulk_insert, db_stats

# ─────────────────────────────────────────────────────────────────────────────
# CORPUS
# Each entry: id, category, subcategory, title, problem, tags, code, notes,
#             source, difficulty
# ─────────────────────────────────────────────────────────────────────────────

ENTRIES = [

# ═══════════════════════════════════════════════════════════════════════════
# CATEGORY: pandas  —  core DataFrame operations
# ═══════════════════════════════════════════════════════════════════════════

{
    "id": "pd-groupby-sum-001",
    "category": "pandas", "subcategory": "groupby",
    "title": "Group by one column and sum numeric columns",
    "problem": "Aggregate a table by a category column and get totals for all numeric columns",
    "tags": ["groupby", "sum", "aggregate", "total", "category"],
    "code": """\
# Basic groupby sum — works on any sheet
df = read_sheet(sheet_name)
result = df.groupby("Category_Column")[["Numeric_Col1","Numeric_Col2"]].sum()
print(result.to_string())
""",
    "notes": "Replace Category_Column and Numeric_Col1/Col2 with actual headers from the graph. Hidden columns are excluded by default.",
    "source": "Python for Data Analysis — Wes McKinney, Ch.10",
    "difficulty": "beginner",
},

{
    "id": "pd-groupby-multi-001",
    "category": "pandas", "subcategory": "groupby",
    "title": "Group by multiple columns with multiple aggregations",
    "problem": "Get sum, mean, and count for each combination of two category columns",
    "tags": ["groupby", "agg", "sum", "mean", "count", "multi-column"],
    "code": """\
df = read_sheet(sheet_name)
result = df.groupby(["Col_A","Col_B"]).agg(
    total   = ("Numeric_Col","sum"),
    average = ("Numeric_Col","mean"),
    count   = ("Numeric_Col","count"),
).reset_index()
print(result.to_string())
""",
    "notes": "Named aggregations (pandas >=0.25) are cleaner than dict-of-lists.",
    "source": "Python for Data Analysis — Wes McKinney, Ch.10",
    "difficulty": "intermediate",
},

{
    "id": "pd-rolling-avg-001",
    "category": "pandas", "subcategory": "window",
    "title": "Rolling N-period moving average",
    "problem": "Smooth a time series with a rolling window average (e.g. 3-month, 12-month)",
    "tags": ["rolling", "moving average", "window", "time-series", "smooth"],
    "code": """\
df = read_sheet(sheet_name)
df = df.sort_values("Date_Column")   # must be sorted first
n = 3   # window size — change to 12 for annual rolling
df["rolling_avg"] = df["Value_Column"].rolling(window=n, min_periods=1).mean()
print(df[["Date_Column","Value_Column","rolling_avg"]].tail(20).to_string())
""",
    "notes": "min_periods=1 avoids NaN at the start. For percentage columns, values are 0-1 — multiply by 100 for display.",
    "source": "Python for Data Analysis — Wes McKinney, Ch.11",
    "difficulty": "intermediate",
},

{
    "id": "pd-running-total-001",
    "category": "pandas", "subcategory": "window",
    "title": "Cumulative sum (running total)",
    "problem": "Add a running total column to a DataFrame",
    "tags": ["cumsum", "running total", "cumulative", "ytd", "year-to-date"],
    "code": """\
df = read_sheet(sheet_name)
df["running_total"] = df["Amount_Column"].cumsum()
print(df[["Amount_Column","running_total"]].to_string())
""",
    "notes": "For running total within a group (e.g. per department): df['running_total'] = df.groupby('Dept')['Amount'].cumsum()",
    "source": "Pandas documentation — GroupBy.cumsum",
    "difficulty": "beginner",
},

{
    "id": "pd-pct-change-001",
    "category": "pandas", "subcategory": "window",
    "title": "Period-over-period percentage change",
    "problem": "Calculate month-over-month or year-over-year growth rate",
    "tags": ["pct_change", "growth", "yoy", "mom", "percentage change"],
    "code": """\
df = read_sheet(sheet_name)
df = df.sort_values("Date_Column")
df["pct_change"] = df["Value_Column"].pct_change() * 100
df["pct_change"] = df["pct_change"].round(2)
print(df[["Date_Column","Value_Column","pct_change"]].to_string())
""",
    "notes": "First row will be NaN (no previous period). Use .fillna(0) if needed.",
    "source": "Pandas documentation — DataFrame.pct_change",
    "difficulty": "beginner",
},

{
    "id": "pd-pivot-table-001",
    "category": "pandas", "subcategory": "pivot",
    "title": "Pivot table: rows x columns with aggregated values",
    "problem": "Create a cross-tabulation showing totals at each row/column intersection",
    "tags": ["pivot", "crosstab", "pivot_table", "matrix", "cross-tab"],
    "code": """\
df = read_sheet(sheet_name)
pivot = df.pivot_table(
    values     = "Numeric_Col",
    index      = "Row_Category",
    columns    = "Col_Category",
    aggfunc    = "sum",
    fill_value = 0,
    margins    = True,   # adds row/column totals
)
print(pivot.to_string())
path = save_new_workbook(pivot, "Pivot_Result")
print("Saved:", path)
""",
    "notes": "margins=True adds Grand Total row/column. Use aggfunc='mean' for averages.",
    "source": "Python for Data Analysis — Wes McKinney, Ch.10",
    "difficulty": "intermediate",
},

{
    "id": "pd-filter-above-avg-001",
    "category": "pandas", "subcategory": "filter",
    "title": "Filter rows where value is above/below column average",
    "problem": "Find rows where a numeric column is significantly above or below average",
    "tags": ["filter", "above average", "outlier", "threshold", "boolean mask"],
    "code": """\
df = read_sheet(sheet_name)
col = "Numeric_Column"
mean_val = df[col].mean()
std_val  = df[col].std()
above_avg = df[df[col] > mean_val]
outliers  = df[(df[col] > mean_val + 2*std_val) | (df[col] < mean_val - 2*std_val)]
print(f"Mean: {mean_val:,.2f}  |  Rows above mean: {len(above_avg)}")
print(f"Outliers (>2 std dev): {len(outliers)}")
print(outliers.to_string())
""",
    "notes": "2 standard deviations covers ~95% of a normal distribution. Use 1.5 for a looser filter.",
    "source": "Python Data Science Handbook — Jake VanderPlas, Ch.3",
    "difficulty": "intermediate",
},

{
    "id": "pd-pct-col-correct-001",
    "category": "pandas", "subcategory": "formatting",
    "title": "Handle percentage columns stored as decimals (0.75 = 75%)",
    "problem": "Excel percentage columns store values as 0-1 decimals. Display correctly and avoid wrong sums.",
    "tags": ["percentage", "decimal", "format", "display", "0.75", "75%"],
    "code": """\
df = read_sheet(sheet_name)
pct_col = "Utilisation_Pct"   # replace with actual column name from graph

# CORRECT: multiply by 100 for display only
print("Average utilisation:", f"{df[pct_col].mean()*100:.1f}%")
print("Max utilisation:",     f"{df[pct_col].max()*100:.1f}%")

# WRONG: df[pct_col].sum()  — gives 12.3 not 1230%

# For charting: format axis
import matplotlib.pyplot as plt
fig, ax = plt.subplots(figsize=(10,5))
ax.bar(df["Category_Col"].astype(str), df[pct_col])
ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x,_: f"{x*100:.0f}%"))
ax.set_title("Utilisation by Category")
plt.tight_layout()
save_chart("utilisation_chart")
""",
    "notes": "Always check graph: format_meaning='percentage' means the column stores 0-1 decimals. Never sum a percentage column.",
    "source": "ByeByeExcel internal — common data quality pattern",
    "difficulty": "beginner",
},

{
    "id": "pd-currency-format-001",
    "category": "pandas", "subcategory": "formatting",
    "title": "Format large numeric amounts for display (thousands, millions, billions)",
    "problem": "Display large amounts in a readable short form for charts and tables",
    "tags": ["currency", "format", "thousands", "millions", "billions", "display"],
    "code": """\
def format_amount(amount, unit="auto"):
    abs_val = abs(amount)
    if unit == "auto":
        if abs_val >= 1e9:   unit = "B"
        elif abs_val >= 1e6: unit = "M"
        elif abs_val >= 1e3: unit = "K"
        else:                unit = ""
    if unit == "B":   return f"{amount/1e9:,.2f}B"
    elif unit == "M": return f"{amount/1e6:,.2f}M"
    elif unit == "K": return f"{amount/1e3:,.2f}K"
    return f"{amount:,.0f}"

df = read_sheet(sheet_name)
col = "Amount_Column"
print(f"Total:   {format_amount(df[col].sum())}")
print(f"Average: {format_amount(df[col].mean())}")

# Display column scaled to millions
df["amount_m"] = (df[col] / 1e6).round(2)
print(df[["Category_Column","amount_m"]].to_string())
""",
    "notes": "Adjust unit thresholds to match your data scale. For currency symbols, prepend your locale symbol to the return string.",
    "source": "Data formatting best practices",
    "difficulty": "beginner",
},

{
    "id": "pd-date-parse-001",
    "category": "pandas", "subcategory": "date",
    "title": "Parse date columns stored as text or Excel serials",
    "problem": "Date columns may be stored as text (01/04/2026), Excel serial numbers (46000+), or mixed formats",
    "tags": ["date", "parse", "datetime", "serial", "format", "to_datetime"],
    "code": """\
df = read_sheet(sheet_name)
date_col = "Date_Column"

# Try text parsing first
try:
    df["date_parsed"] = pd.to_datetime(df[date_col], dayfirst=True, errors="coerce")
    bad = df["date_parsed"].isna().sum()
    if bad > len(df) * 0.1:  # >10% failed — try serial
        raise ValueError("too many NaT")
except Exception:
    # Excel serial: days since 1899-12-30
    df["date_parsed"] = pd.to_datetime(
        df[date_col], unit="D", origin="1899-12-30", errors="coerce")

df["year"]       = df["date_parsed"].dt.year
df["month"]      = df["date_parsed"].dt.month
df["month_name"] = df["date_parsed"].dt.strftime("%b %Y")
print(df[["date_parsed","year","month"]].head().to_string())
""",
    "notes": "dayfirst=True handles DD/MM/YYYY format. Set dayfirst=False for MM/DD/YYYY (US format). Excel stores dates as integers since 1899-12-30.",
    "source": "Pandas documentation — to_datetime, plus Excel date system spec",
    "difficulty": "intermediate",
},

{
    "id": "pd-merge-sheets-001",
    "category": "pandas", "subcategory": "merge",
    "title": "Merge two sheets on a common key column (VLOOKUP equivalent)",
    "problem": "Join data from two sheets like a VLOOKUP — bring a column from Sheet2 into Sheet1",
    "tags": ["merge", "join", "vlookup", "lookup", "two sheets", "combine"],
    "code": """\
sheets = read_all_sheets()
df1 = sheets["Sheet1_Name"]   # main data
df2 = sheets["Sheet2_Name"]   # lookup table

# Left join — keeps all rows from df1, adds matching data from df2
merged = df1.merge(
    df2[["Key_Column","Value_To_Add"]],
    on  = "Key_Column",
    how = "left",
)
# Rows with no match will have NaN in Value_To_Add
unmatched = merged["Value_To_Add"].isna().sum()
print(f"Rows: {len(merged)} | Unmatched: {unmatched}")
print(merged.head().to_string())
path = save_new_workbook(merged, "Merged_Result")
print("Saved:", path)
""",
    "notes": "Use how='inner' to keep only matching rows. how='left' keeps all from left table.",
    "source": "Python for Data Analysis — Wes McKinney, Ch.8",
    "difficulty": "intermediate",
},

{
    "id": "pd-deduplicate-001",
    "category": "pandas", "subcategory": "dedup",
    "title": "Remove exact duplicate rows, keep first occurrence",
    "problem": "Find and remove identical rows from a DataFrame",
    "tags": ["duplicate", "dedup", "drop_duplicates", "unique", "exact match"],
    "code": """\
df = read_sheet(sheet_name)
n_before = len(df)
df_clean = df.drop_duplicates()
n_after  = len(df_clean)
print(f"Before: {n_before:,} rows  |  After: {n_after:,}  |  Removed: {n_before-n_after:,}")

# See the duplicates
dups = df[df.duplicated(keep=False)]
print(f"\\nDuplicate rows (showing both copies):")
print(dups.head(10).to_string())

path = save_new_workbook(df_clean, "Deduped")
print("Clean data saved:", path)
""",
    "notes": "Use drop_duplicates(subset=['Col_A','Col_B']) to only consider specific columns for dedup.",
    "source": "Pandas documentation — DataFrame.drop_duplicates",
    "difficulty": "beginner",
},

{
    "id": "pd-rank-top-n-001",
    "category": "pandas", "subcategory": "filter",
    "title": "Get top N rows by a numeric column",
    "problem": "Find the top 10 or bottom 10 values in a column, optionally within groups",
    "tags": ["top", "nlargest", "nsmallest", "rank", "top 10", "highest"],
    "code": """\
df = read_sheet(sheet_name)
col = "Numeric_Column"
n   = 10

top_n    = df.nlargest(n, col)
bottom_n = df.nsmallest(n, col)

print(f"Top {n} by {col}:")
print(top_n.to_string())

# Top N within each group
top_per_group = (
    df.sort_values(col, ascending=False)
      .groupby("Category_Col")
      .head(n)
)
print(f"\\nTop {n} per category:")
print(top_per_group.to_string())
""",
    "notes": "nlargest/nsmallest are faster than sort_values().head() for large DataFrames.",
    "source": "Pandas documentation — DataFrame.nlargest",
    "difficulty": "beginner",
},

{
    "id": "pd-crosstab-001",
    "category": "pandas", "subcategory": "pivot",
    "title": "Cross-tabulation: count occurrences of two categorical columns",
    "problem": "Count how many rows have each combination of two category columns",
    "tags": ["crosstab", "frequency", "count", "cross tab", "matrix count"],
    "code": """\
df = read_sheet(sheet_name)
ct = pd.crosstab(df["Row_Category"], df["Col_Category"], margins=True)
print(ct.to_string())

# As percentages
ct_pct = pd.crosstab(df["Row_Category"], df["Col_Category"], normalize="index") * 100
ct_pct = ct_pct.round(1)
print("\\nAs % of row total:")
print(ct_pct.to_string())
""",
    "notes": "normalize='index' shows row %, normalize='columns' shows column %, normalize=True shows total %.",
    "source": "Python for Data Analysis — Wes McKinney, Ch.10",
    "difficulty": "intermediate",
},

{
    "id": "pd-missing-fill-001",
    "category": "pandas", "subcategory": "cleaning",
    "title": "Fill missing values with column mean, median, or forward fill",
    "problem": "Replace NaN values in a column with a sensible default",
    "tags": ["fillna", "missing", "NaN", "impute", "forward fill", "mean"],
    "code": """\
df = read_sheet(sheet_name)
col = "Column_With_Nulls"

# Strategy 1: fill with column mean (numeric only)
df[col + "_filled_mean"]   = df[col].fillna(df[col].mean())

# Strategy 2: fill with median (robust to outliers)
df[col + "_filled_median"] = df[col].fillna(df[col].median())

# Strategy 3: forward fill (carry previous value forward — good for time series)
df[col + "_ffill"]         = df[col].ffill()

# Strategy 4: fill text columns with 'Unknown'
text_col = "Text_Column"
df[text_col] = df[text_col].fillna("Unknown")

print(f"Original NaN count: {df[col].isna().sum()}")
""",
    "notes": "For financial data, median is usually safer than mean (outlier-resistant). For time series, ffill preserves trend.",
    "source": "Python Data Science Handbook — Jake VanderPlas, Ch.3",
    "difficulty": "beginner",
},

# ═══════════════════════════════════════════════════════════════════════════
# CATEGORY: analysis  —  statistical + business intelligence patterns
# ═══════════════════════════════════════════════════════════════════════════

{
    "id": "analysis-pareto-001",
    "category": "analysis", "subcategory": "distribution",
    "title": "Pareto analysis: 80/20 rule — which categories drive 80% of total",
    "problem": "Find which categories account for 80% of a total metric (the vital few)",
    "tags": ["pareto", "80/20", "cumulative", "vital few", "ABC analysis"],
    "code": """\
df = read_sheet(sheet_name)
col = "Amount_Column"
cat = "Category_Column"

summary = df.groupby(cat)[col].sum().reset_index()
summary = summary.sort_values(col, ascending=False)
summary["cumulative_pct"] = summary[col].cumsum() / summary[col].sum() * 100
summary["pareto_80"] = summary["cumulative_pct"] <= 80

print("Categories driving 80% of total:")
print(summary[summary["pareto_80"]].to_string())
print(f"\\n{summary['pareto_80'].sum()} out of {len(summary)} categories = 80% of total")

# Chart
import matplotlib.pyplot as plt
fig, ax1 = plt.subplots(figsize=(12,6))
ax2 = ax1.twinx()
ax1.bar(range(len(summary)), summary[col], color="#4472C4", alpha=0.7)
ax2.plot(range(len(summary)), summary["cumulative_pct"], color="#C0392B",
         linewidth=2, marker="o", markersize=4)
ax2.axhline(80, color="#C0392B", linestyle="--", alpha=0.5)
ax1.set_xticks(range(len(summary)))
ax1.set_xticklabels(summary[cat].astype(str), rotation=45, ha="right")
ax1.set_ylabel(col); ax2.set_ylabel("Cumulative %")
plt.title("Pareto Analysis"); plt.tight_layout()
save_chart("pareto_chart")
""",
    "notes": "Pareto principle: in most datasets, 20% of categories drive 80% of the metric.",
    "source": "Data Analysis with Python — multiple sources",
    "difficulty": "intermediate",
},

{
    "id": "analysis-variance-001",
    "category": "analysis", "subcategory": "financial",
    "title": "Budget vs Actual variance analysis",
    "problem": "Compare budget and actual amounts, compute variance and % variance",
    "tags": ["variance", "budget", "actual", "comparison", "overspend", "underspend"],
    "code": """\
df = read_sheet(sheet_name)
budget_col = "Budget_Column"
actual_col = "Actual_Column"
cat_col    = "Category_Column"

df["variance"]     = df[actual_col] - df[budget_col]
df["variance_pct"] = (df["variance"] / df[budget_col].replace(0, float("nan"))) * 100
df["status"]       = df["variance"].apply(
    lambda x: "Overspent" if x > 0 else ("Underspent" if x < 0 else "On Budget"))

summary = df.groupby(cat_col).agg(
    budget   = (budget_col, "sum"),
    actual   = (actual_col, "sum"),
    variance = ("variance", "sum"),
).reset_index()
summary["variance_pct"] = (summary["variance"] / summary["budget"] * 100).round(1)

print(summary.to_string())
print(f"\\nTotal overspent: {df[df['variance']>0]['variance'].sum():,.0f}")
path = save_new_workbook(summary, "Budget_Variance")
print("Saved:", path)
""",
    "notes": "Positive variance = overspent (actual > budget). Convention varies — confirm with your stakeholder.",
    "source": "Financial analysis best practices",
    "difficulty": "intermediate",
},

{
    "id": "analysis-utilisation-001",
    "category": "analysis", "subcategory": "financial",
    "title": "Expenditure utilisation rate by category",
    "problem": "Calculate what percentage of allocated budget has been spent, flag low and high utilisation",
    "tags": ["utilisation", "expenditure", "budget", "percentage spent", "absorption"],
    "code": """\
df = read_sheet(sheet_name)
budget_col = "Budget_Column"
actual_col = "Actual_Column"
cat_col    = "Category_Column"

df["utilisation_pct"] = (df[actual_col] / df[budget_col].replace(0, float("nan"))).clip(0, 2)
df["band"] = df["utilisation_pct"].apply(
    lambda x: "Critical (<25%)" if x < 0.25
    else "Low (25-50%)"        if x < 0.50
    else "Moderate (50-75%)"   if x < 0.75
    else "Good (75-90%)"       if x < 0.90
    else "High (>90%)")

print(df[[cat_col,"utilisation_pct","band"]].sort_values("utilisation_pct").to_string())
print("\\nDistribution:")
print(df["band"].value_counts().to_string())

# Chart
import matplotlib.pyplot as plt
fig, ax = plt.subplots(figsize=(12,6))
colours = {"Critical (<25%)":"#c0392b","Low (25-50%)":"#e67e22",
           "Moderate (50-75%)":"#f1c40f","Good (75-90%)":"#27ae60",
           "High (>90%)":"#2980b9"}
df_sorted = df.sort_values("utilisation_pct")
ax.barh(df_sorted[cat_col].astype(str),
        df_sorted["utilisation_pct"] * 100,
        color=[colours[b] for b in df_sorted["band"]])
ax.axvline(75, color="#27ae60", linestyle="--", linewidth=1.5, label="75% target")
ax.set_xlabel("Utilisation %")
ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x,_: f"{x:.0f}%"))
ax.legend(); plt.tight_layout()
save_chart("utilisation_chart")
""",
    "notes": "clip(0,2) prevents negative or >200% values from distorting analysis. Adjust band thresholds to match your organisation's policy.",
    "source": "Financial management best practices",
    "difficulty": "intermediate",
},

{
    "id": "analysis-rank-percentile-001",
    "category": "analysis", "subcategory": "distribution",
    "title": "Add rank and percentile columns to a dataset",
    "problem": "Rank rows by a metric and compute their percentile position",
    "tags": ["rank", "percentile", "quartile", "position", "benchmark"],
    "code": """\
df = read_sheet(sheet_name)
col = "Numeric_Column"

df["rank"]       = df[col].rank(method="dense", ascending=False).astype(int)
df["percentile"] = df[col].rank(pct=True) * 100
df["quartile"]   = pd.qcut(df[col], q=4,
                            labels=["Q1 (bottom 25%)","Q2","Q3","Q4 (top 25%)"])

print(df[[col,"rank","percentile","quartile"]]
      .sort_values("rank").head(20).to_string())
""",
    "notes": "method='dense' gives consecutive ranks without gaps. method='min' gives standard competition ranks.",
    "source": "Pandas documentation — Series.rank",
    "difficulty": "intermediate",
},

# ═══════════════════════════════════════════════════════════════════════════
# CATEGORY: sql  —  SQLite queries for large files
# ═══════════════════════════════════════════════════════════════════════════

{
    "id": "sql-groupby-001",
    "category": "sql", "subcategory": "groupby",
    "title": "SQL GROUP BY with SUM and COUNT on SQLite database",
    "problem": "Aggregate a large table by category without loading into RAM",
    "tags": ["sql", "group by", "sum", "count", "sqlite", "large file"],
    "code": """\
import sqlite3, pandas as pd
conn = sqlite3.connect(GRAPH["db_path"])
table = list(GRAPH["sheet_tables"].values())[0]["table"]   # first sheet table
result = pd.read_sql(f\"\"\"
    SELECT
        _c_district         AS region,
        COUNT(*)            AS total_rows,
        SUM(CAST(_c_amount AS REAL)) AS total_amount,
        AVG(CAST(_c_amount AS REAL)) AS avg_amount
    FROM [{table}]
    WHERE _c_amount IS NOT NULL
    GROUP BY _c_district
    ORDER BY total_amount DESC
\"\"\", conn)
conn.close()
print(result.to_string())
""",
    "notes": "Canonical columns are prefixed _c_ (e.g. _c_district, _c_amount). Use GRAPH['sheet_tables'] to get the correct table name per sheet. Always CAST to REAL for numeric operations in SQLite.",
    "source": "SQLite documentation + ByeByeExcel schema",
    "difficulty": "intermediate",
},

{
    "id": "sql-shared-id-001",
    "category": "sql", "subcategory": "anomaly",
    "title": "SQL: find ID numbers shared across multiple names",
    "problem": "Detect records where the same ID number appears under different names — a strong data quality signal",
    "tags": ["sql", "id", "shared", "duplicate", "group by having", "data quality"],
    "code": """\
import sqlite3, pandas as pd
conn = sqlite3.connect(GRAPH["db_path"])
table = list(GRAPH["sheet_tables"].values())[0]["table"]
result = pd.read_sql(f\"\"\"
    SELECT
        _c_id_field                         AS id_value,
        COUNT(DISTINCT _c_name)             AS distinct_names,
        GROUP_CONCAT(DISTINCT _c_name)      AS names_list,
        COUNT(*)                            AS total_rows
    FROM [{table}]
    WHERE _c_id_field IS NOT NULL
    GROUP BY _c_id_field
    HAVING COUNT(DISTINCT _c_name) > 1
    ORDER BY distinct_names DESC
    LIMIT 100
\"\"\", conn)
conn.close()
print(f"Shared ID entries: {{len(result)}}")
print(result.to_string())
""",
    "notes": "Replace _c_id_field with the canonical column for your ID field (e.g. _c_aadhaar, _c_ssn, _c_employee_id). Requires excel_ingest_large first for large files.",
    "source": "Data quality anomaly detection patterns",
    "difficulty": "advanced",
},

{
    "id": "sql-location-density-001",
    "category": "sql", "subcategory": "anomaly",
    "title": "SQL: detect locations with anomalously high record counts",
    "problem": "Flag locations where record count is more than 3 standard deviations above the mean — statistical outlier detection",
    "tags": ["sql", "location", "density", "anomaly", "statistical", "zscore", "outlier"],
    "code": """\
import sqlite3, pandas as pd
conn = sqlite3.connect(GRAPH["db_path"])
table = list(GRAPH["sheet_tables"].values())[0]["table"]
result = pd.read_sql(f\"\"\"
    WITH location_counts AS (
        SELECT _c_location AS location, COUNT(*) AS cnt
        FROM [{table}]
        WHERE _c_location IS NOT NULL
        GROUP BY _c_location
    ),
    stats AS (
        SELECT
            AVG(cnt)                          AS mean_cnt,
            AVG(cnt*cnt) - AVG(cnt)*AVG(cnt)  AS var_cnt
        FROM location_counts
    )
    SELECT
        v.location,
        v.cnt,
        ROUND(s.mean_cnt, 1) AS mean_cnt,
        ROUND((v.cnt - s.mean_cnt) / MAX(1.0, SQRT(s.var_cnt)), 2) AS z_score
    FROM location_counts v
    CROSS JOIN stats s
    WHERE s.var_cnt > 0
      AND (v.cnt - s.mean_cnt) / MAX(1.0, SQRT(s.var_cnt)) > 3.0
    ORDER BY z_score DESC
\"\"\", conn)
conn.close()
print(f"Anomalous locations: {{len(result)}}")
print(result.to_string())
""",
    "notes": "Replace _c_location with your location column (e.g. _c_city, _c_region, _c_district). Z-score > 3 means more than 3 standard deviations above mean. Adjust threshold as needed.",
    "source": "Statistical anomaly detection — ByeByeExcel patterns",
    "difficulty": "advanced",
},

# ═══════════════════════════════════════════════════════════════════════════
# CATEGORY: visualisation  —  matplotlib chart patterns
# ═══════════════════════════════════════════════════════════════════════════

{
    "id": "viz-bar-grouped-001",
    "category": "visualisation", "subcategory": "bar",
    "title": "Grouped bar chart: compare two metrics side by side",
    "problem": "Show Budget vs Actual (or any two metrics) as grouped bars for each category",
    "tags": ["bar chart", "grouped", "compare", "side by side", "budget vs actual"],
    "code": """\
import matplotlib.pyplot as plt
import numpy as np
df = read_sheet(sheet_name)

categories = df["Category_Column"].astype(str).tolist()
metric_a   = df["Budget_Column"].tolist()
metric_b   = df["Actual_Column"].tolist()

x     = np.arange(len(categories))
width = 0.35

fig, ax = plt.subplots(figsize=(max(10, len(categories)*0.8), 6))
ax.bar(x - width/2, metric_a, width, label="Budget", color="#4472C4")
ax.bar(x + width/2, metric_b, width, label="Actual", color="#70AD47")

ax.set_xticks(x)
ax.set_xticklabels(categories, rotation=45, ha="right")
ax.set_ylabel("Amount")
ax.set_title("Budget vs Actual by Category")
ax.legend()
plt.tight_layout()
save_chart("budget_vs_actual")
""",
    "notes": "Adjust figure width: len(categories)*0.8 keeps bars readable at any size.",
    "source": "Matplotlib documentation",
    "difficulty": "intermediate",
},

{
    "id": "viz-waterfall-001",
    "category": "visualisation", "subcategory": "chart",
    "title": "Waterfall chart showing incremental changes",
    "problem": "Show how a total builds up from components, or how it changed from one period to another",
    "tags": ["waterfall", "bridge chart", "incremental", "variance", "breakdown"],
    "code": """\
import matplotlib.pyplot as plt

df = read_sheet(sheet_name)
labels = df["Category_Column"].astype(str).tolist()
values = df["Amount_Column"].tolist()

# Compute running total and bar bottoms
running = [0]
bottoms = []
for v in values:
    bottoms.append(running[-1] if v >= 0 else running[-1] + v)
    running.append(running[-1] + v)

colours = ["#70AD47" if v >= 0 else "#C0392B" for v in values]

fig, ax = plt.subplots(figsize=(max(10, len(labels)*1.2), 6))
ax.bar(range(len(labels)), [abs(v) for v in values],
       bottom=bottoms, color=colours, edgecolor="white", linewidth=0.5)

for i, (v, b) in enumerate(zip(values, bottoms)):
    ax.text(i, b + abs(v)/2, f"{v:+,.0f}",
            ha="center", va="center", fontsize=9, color="white", fontweight="bold")

ax.set_xticks(range(len(labels)))
ax.set_xticklabels(labels, rotation=45, ha="right")
ax.set_title("Waterfall Chart")
plt.tight_layout()
save_chart("waterfall_chart")
""",
    "notes": "Green bars = positive contribution, Red = negative. Works well for variance explanation and build-up charts.",
    "source": "Data visualisation best practices",
    "difficulty": "advanced",
},

{
    "id": "viz-heatmap-001",
    "category": "visualisation", "subcategory": "heatmap",
    "title": "Heatmap for cross-tabulation of two categorical dimensions",
    "problem": "Visualise a matrix of values as a colour grid — useful for spotting patterns across two dimensions",
    "tags": ["heatmap", "matrix", "cross-tab", "colour grid", "intensity"],
    "code": """\
import matplotlib.pyplot as plt

df = read_sheet(sheet_name)
pivot = df.pivot_table(
    values   = "Numeric_Column",
    index    = "Row_Category",
    columns  = "Col_Category",
    aggfunc  = "mean",
    fill_value = 0,
)

fig, ax = plt.subplots(figsize=(max(8, len(pivot.columns)*1.2),
                                max(6, len(pivot)*0.5)))
im = ax.imshow(pivot.values, cmap="RdYlGn", aspect="auto")

ax.set_xticks(range(len(pivot.columns)))
ax.set_xticklabels(pivot.columns.astype(str), rotation=45, ha="right", fontsize=9)
ax.set_yticks(range(len(pivot.index)))
ax.set_yticklabels(pivot.index.astype(str), fontsize=9)

plt.colorbar(im, ax=ax)
ax.set_title("Heatmap")
plt.tight_layout()
save_chart("heatmap")
""",
    "notes": "RdYlGn = Red (low) -> Yellow (mid) -> Green (high). Good for utilisation. Use 'Blues' for volume.",
    "source": "Matplotlib documentation + data visualisation best practices",
    "difficulty": "advanced",
},

# ═══════════════════════════════════════════════════════════════════════════
# CATEGORY: excel  —  openpyxl workbook manipulation
# ═══════════════════════════════════════════════════════════════════════════

{
    "id": "excel-conditional-format-001",
    "category": "excel", "subcategory": "formatting",
    "title": "Apply conditional formatting: colour cells by value threshold",
    "problem": "Highlight cells in a column red/yellow/green based on their value",
    "tags": ["conditional formatting", "highlight", "colour", "threshold", "openpyxl"],
    "code": """\
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.load_workbook(AI_WORKBOOK)
ws = wb["Sheet_Name"]
last_row = ws.max_row

red    = PatternFill("solid", fgColor="FFFFC7CE")
yellow = PatternFill("solid", fgColor="FFFFEB9C")
green  = PatternFill("solid", fgColor="FFC6EFCE")

# Column K = percentage column; thresholds at 0.75 and 0.90
ws.conditional_formatting.add(f"K2:K{last_row}",
    CellIsRule(operator="lessThan",   formula=["0.75"], fill=yellow))
ws.conditional_formatting.add(f"K2:K{last_row}",
    CellIsRule(operator="greaterThan",formula=["0.90"], fill=red))
ws.conditional_formatting.add(f"K2:K{last_row}",
    CellIsRule(operator="between",    formula=["0.75","0.90"], fill=green))

wb.save(AI_WORKBOOK)
print(f"Conditional formatting applied to K2:K{last_row}")
""",
    "notes": "Always write to AI_WORKBOOK, never the original. Column letter must match the graph. Adjust thresholds for your use case.",
    "source": "openpyxl documentation",
    "difficulty": "intermediate",
},

{
    "id": "excel-write-formula-001",
    "category": "excel", "subcategory": "formula",
    "title": "Write SUM / AVERAGE formulas to a column using openpyxl",
    "problem": "Programmatically add formulas to a column in the ai_workbook",
    "tags": ["formula", "write", "openpyxl", "SUM", "AVERAGE", "programmatic"],
    "code": """\
import openpyxl
wb = openpyxl.load_workbook(AI_WORKBOOK)
ws = wb["Sheet_Name"]

# Add a Total row below data
last_data_row = ws.max_row
total_row     = last_data_row + 2

ws.cell(total_row, 1, "TOTAL")
for col_letter in ["E","F","G","H","J"]:   # numeric columns from graph
    ws.cell(total_row, openpyxl.utils.column_index_from_string(col_letter),
            f"=SUM({col_letter}2:{col_letter}{last_data_row})")

wb.save(AI_WORKBOOK)
print(f"Totals written to row {total_row}")
""",
    "notes": "Always use column letters from the graph. Never hardcode row numbers — use max_row.",
    "source": "openpyxl documentation",
    "difficulty": "intermediate",
},

# ═══════════════════════════════════════════════════════════════════════════
# CATEGORY: dedup  —  patterns for fuzzy matching
# ═══════════════════════════════════════════════════════════════════════════

{
    "id": "dedup-phonetic-name-001",
    "category": "dedup", "subcategory": "names",
    "title": "Deduplicate names with phonetic variants using geographic blocking",
    "problem": "Find near-duplicate records where names differ due to phonetic spelling variants across data entry operators",
    "tags": ["dedup", "names", "phonetic", "blocking", "fuzzy", "geographic", "run_code"],
    "code": """\
# Fast approach: block by location + name prefix, then compare with SequenceMatcher
# Avoids O(N^2) on large files. Works well when location columns are available.
import pandas as pd
from difflib import SequenceMatcher

df = read_sheet(sheet_name)

# Blocking key: location columns + first 3 chars of name
# Adjust column names to match your graph
name_col     = "Name_Column"
location_col = "City_Column"     # or Region, District, etc.
parent_col   = "Parent_Name_Column"  # secondary validation column

df["_block"] = (
    df[location_col].astype(str) + "_" +
    df[name_col].astype(str).str[:3].str.upper()
)

NAME_THRESH   = 0.80
PARENT_THRESH = 0.70
matches = []

for _, group in df.groupby("_block"):
    if len(group) < 2:
        continue
    records = group.to_dict("records")
    for i in range(len(records)):
        for j in range(i + 1, len(records)):
            r1, r2 = records[i], records[j]
            name_sim = SequenceMatcher(
                None, str(r1[name_col]), str(r2[name_col])).ratio()
            if name_sim < NAME_THRESH:
                continue
            parent_sim = SequenceMatcher(
                None, str(r1[parent_col]), str(r2[parent_col])).ratio()
            if parent_sim < PARENT_THRESH:
                continue
            matches.append({
                "Name_A":       r1[name_col],
                "Name_B":       r2[name_col],
                "Parent_A":     r1[parent_col],
                "Parent_B":     r2[parent_col],
                "Location":     r1[location_col],
                "Name_Sim":     round(name_sim, 3),
                "Parent_Sim":   round(parent_sim, 3),
            })

if matches:
    result = pd.DataFrame(matches).sort_values("Name_Sim", ascending=False)
    print(f"Found {len(result):,} potential duplicate pairs")
    if len(result) > 40:
        path = save_new_workbook(result, "Fuzzy_Matches")
        print(f"Saved to: {path}")
    else:
        print(result.to_string())
else:
    print("No near matches found")
""",
    "notes": "Location blocking (city/region + name prefix) is critical for performance on large files. Without it, comparison is O(N^2). The secondary column (parent/employer/department) validates that high name similarity is a true match, not coincidence.",
    "source": "ByeByeExcel dedup patterns",
    "difficulty": "intermediate",
},

{
    "id": "dedup-multi-column-001",
    "category": "dedup", "subcategory": "strategy",
    "title": "Multi-column weighted dedup: name + ID + amount",
    "problem": "Use multiple columns as evidence for deduplication, with different weights per column",
    "tags": ["dedup", "multi-column", "weights", "composite score", "evidence"],
    "code": """\
# Vendor/contractor dedup: Company name (high weight) + ID (medium) + amount (low)
result = excel_fuzzy_dedup(
    path              = AI_WORKBOOK,
    columns           = "Company_Name,ID_Number,Contract_Amount",
    normalisers       = "generic_text,code_id,numeric_range",
    weights           = "0.65,0.25,0.10",
    normaliser_params = '{"numeric_range": {"tolerance": 0.05}}',
    high_threshold    = 0.85,
    review_threshold  = 0.70,
    sheet             = "Sheet_Name",
)
""",
    "notes": "code_id normaliser strips spaces and separators from ID numbers. numeric_range 5% tolerance handles minor amount rounding differences. Adjust weights so the most reliable field has the highest weight.",
    "source": "ByeByeExcel normaliser patterns",
    "difficulty": "advanced",
},

]


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def seed(db_path: str = None):
    path = init_knowledge_db(db_path)
    result = bulk_insert(ENTRIES, path)
    print(f"Seed complete -> {path}")
    print(f"  Inserted: {result['inserted']}")
    print(f"  Skipped (already exist): {result['skipped']}")
    print(f"  Total: {result['total']}")
    stats = db_stats(path)
    print(f"\nDB stats:")
    for cat, cnt in stats.get("by_category", {}).items():
        print(f"  {cat}: {cnt} entries")
    return path


if __name__ == "__main__":
    import argparse, json
    parser = argparse.ArgumentParser()
    parser.add_argument("--db",    default=None, help="Path to knowledge DB")
    parser.add_argument("--stats", action="store_true", help="Show stats only")
    args = parser.parse_args()

    if args.stats:
        from knowledge import db_stats
        p = args.db or str(Path(__file__).parent / "excel_knowledge.db")
        print(json.dumps(db_stats(p), indent=2))
    else:
        seed(args.db)