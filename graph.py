"""
graph.py — Workbook structural and formatting inspection.

Extracted from v1 server.py into a standalone module.
Used by server.py for excel_init and excel_inspect.

Two passes (both read-only after the copy is made):
  Pass 1 — openpyxl read_only=True:  formulas, cross-sheet refs, named ranges,
            headers, column types, sample rows, tables, VBA. Memory-safe.
  Pass 2 — openpyxl read_only=False: formatting (number_format, fill colour,
            hidden columns). Opens ai_workbook, reads rows 1-2 only, closes.

Returns a graph dict suitable for JSON serialisation and LLM context.
"""

import re, json, zipfile, datetime
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# COLOUR → SEMANTIC MEANING (standard financial model convention)
# ─────────────────────────────────────────────────────────────────────────────

_COLOUR_MEANINGS = {
    "FF0000FF": "input_hardcoded", "FF4472C4": "input_hardcoded",
    "FF9DC3E6": "input_hardcoded", "FFDAE3F3": "input_hardcoded",
    "FF00B050": "link_cross_sheet","FF70AD47": "link_cross_sheet",
    "FFE2EFDA": "link_cross_sheet",
    "FFFFFF00": "assumption",      "FFFFEB9C": "assumption",
    "FFFFC000": "assumption",      "FFFFCC00": "assumption",
    "FFFF0000": "external_link_or_flag",
    "FFD9D9D9": "output_display",  "FFF2F2F2": "output_display",
    "FFE7E6E6": "output_display",
}

_NUMBER_FORMAT_MEANINGS = {
    "0%": "percentage", "0.0%": "percentage", "0.00%": "percentage",
    "$":  "currency_usd", "₹": "currency_inr",
    "£":  "currency_gbp", "€": "currency_eur",
    "#,##0": "number_thousands",
    "DD/MM": "date", "MM/DD": "date", "YYYY": "date",
    "YY": "date", "h:mm": "time",
    "General": "general", "@": "text_forced",
}


def _fmt_meaning(fmt: str) -> str:
    if not fmt or fmt == "General": return "general"
    fu = fmt.upper()
    for pat, label in _NUMBER_FORMAT_MEANINGS.items():
        if pat.upper() in fu: return label
    return "custom"


def _colour_meaning(rgb: str) -> str:
    if not rgb or rgb in ("00000000","FFFFFFFF","FF000000"): return "none"
    return _COLOUR_MEANINGS.get(rgb.upper(), f"custom:{rgb}")


def _col_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _infer_dtype(values: list) -> str:
    non_null = [v for v in values if v is not None]
    if not non_null: return "empty"
    if all(isinstance(v, bool) for v in non_null): return "boolean"
    if all(isinstance(v, (int, float)) for v in non_null): return "numeric"
    if all(hasattr(v, "year") for v in non_null): return "date"
    return "text"


def _sheet_refs(formula: str) -> list:
    return list(set(re.findall(r"'?([A-Za-z0-9_\- ]+)'?!", formula)))


def _extract_vba(path: str) -> dict:
    result = {}
    p = Path(path)
    if p.suffix.lower() not in (".xlsm", ".xlam"):
        return result
    try:
        from oletools.olevba import VBA_Parser
        vba = VBA_Parser(str(p))
        if vba.detect_vba_macros():
            for (_, mod, _, code) in vba.extract_macros():
                result[mod] = code
        vba.close()
        return result
    except ImportError:
        pass
    try:
        with zipfile.ZipFile(p) as z:
            if "xl/vbaProject.bin" in z.namelist():
                raw = z.read("xl/vbaProject.bin").decode("latin-1", errors="replace")
                subs = re.findall(
                    r"(?:Sub|Function)\s+\w+.*?End\s+(?:Sub|Function)",
                    raw, re.DOTALL)
                for i, s in enumerate(subs):
                    result[f"Module_{i+1}"] = "".join(
                        c if c.isprintable() else " " for c in s)[:3000]
    except Exception:
        pass
    return result


# ─────────────────────────────────────────────────────────────────────────────
# PASS 1: STRUCTURAL (read_only=True, memory-safe)
# ─────────────────────────────────────────────────────────────────────────────

def inspect_structure(path: str) -> dict:
    wb_f = openpyxl.load_workbook(path, data_only=False, read_only=True)
    wb_v = openpyxl.load_workbook(path, data_only=True,  read_only=True)

    g = {
        "path":             str(Path(path).resolve()),
        "created_at":       datetime.datetime.now().isoformat(timespec="seconds"),
        "sheet_names":      wb_f.sheetnames,
        "named_ranges":     {},
        "cross_sheet_refs": {},
        "sheets":           {},
        "vba":              {},
    }

    for name, defn in wb_f.defined_names.items():
        dests = []
        try:
            for title, coord in defn.destinations:
                dests.append(f"{title}!{coord}")
        except Exception:
            dests.append(str(getattr(defn, "attr_text", defn)))
        g["named_ranges"][name] = dests

    cross = defaultdict(set)

    for sname in wb_f.sheetnames:
        ws_f = wb_f[sname]
        ws_v = wb_v[sname]

        s = {
            "name":         sname,
            "max_row":      ws_f.max_row or 0,
            "max_col":      ws_f.max_column or 0,
            "headers":      [],
            "column_types": {},
            "column_formats": {},     # filled by pass 2
            "sample_rows":  [],
            "formulas":     {"total": 0, "patterns": []},
            "merged_cells": [],
            "tables":       [],
        }

        # Headers from row 1
        first_row = next(ws_v.iter_rows(min_row=1, max_row=1, values_only=True), ())
        s["headers"] = list(first_row)

        # Column type inference from sample rows 2-11
        col_samples = defaultdict(list)
        for row in ws_v.iter_rows(min_row=2, max_row=11, values_only=True):
            for ci, val in enumerate(row):
                col_samples[ci].append(val)
        for ci, hdr in enumerate(s["headers"]):
            label = str(hdr) if hdr is not None else _col_letter(ci + 1)
            s["column_types"][label] = _infer_dtype(col_samples[ci])

        # Sample rows 2-5 (values)
        for row in ws_v.iter_rows(min_row=2, max_row=5, values_only=True):
            rd = {}
            for ci, val in enumerate(row):
                h = s["headers"][ci] if ci < len(s["headers"]) else _col_letter(ci+1)
                rd[str(h) if h else _col_letter(ci+1)] = val
            s["sample_rows"].append(rd)

        # Formula patterns
        patterns: dict = {}
        for row in ws_f.iter_rows(values_only=False):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith("="):
                    s["formulas"]["total"] += 1
                    refs = _sheet_refs(val)
                    for t in refs:
                        cross[sname].add(t)
                    pat = re.sub(r"\$?[A-Z]{1,3}\$?\d+", "§", val)
                    if pat not in patterns:
                        patterns[pat] = {
                            "example_cell": cell.coordinate,
                            "formula":      val,
                            "cross_sheet_refs": refs,
                            "count":        0,
                            "cells":        [],
                        }
                    patterns[pat]["count"] += 1
                    if len(patterns[pat]["cells"]) < 5:
                        patterns[pat]["cells"].append(cell.coordinate)

        s["formulas"]["patterns"] = list(patterns.values())[:40]

        try:
            s["merged_cells"] = [str(r) for r in ws_f.merged_cells.ranges]
        except Exception:
            pass
        try:
            for t in ws_f.tables.values():
                s["tables"].append({
                    "name":    t.displayName,
                    "ref":     t.ref,
                    "columns": [c.name for c in t.tableColumns],
                })
        except Exception:
            pass

        g["sheets"][sname] = s

    g["cross_sheet_refs"] = {k: sorted(v) for k, v in cross.items()}
    g["vba"] = _extract_vba(path)
    wb_f.close()
    wb_v.close()
    return g


# ─────────────────────────────────────────────────────────────────────────────
# PASS 2: FORMATTING (read_only=False, rows 1-2 only)
# ─────────────────────────────────────────────────────────────────────────────

def inspect_formatting(path: str) -> dict:
    """
    Open with read_only=False but read only 2 rows per sheet.
    Returns {sheet_name: {col_header: format_info_dict}}.
    Fast even for 1L+ row files.
    """
    result = {}
    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
    except Exception as e:
        return {"error": str(e)}

    for sname in wb.sheetnames:
        ws = wb[sname]
        sheet_fmt = {}

        headers = [cell.value for cell in ws[1]]

        col_dims = {}
        for col_idx in range(1, (ws.max_column or 0) + 1):
            letter = get_column_letter(col_idx)
            dim = ws.column_dimensions.get(letter)
            col_dims[col_idx] = {
                "hidden": bool(dim and dim.hidden),
                "width":  round(dim.width, 1) if dim and dim.width else None,
            }

        row2_cells = list(ws.iter_rows(min_row=2, max_row=2, values_only=False))
        row2 = row2_cells[0] if row2_cells else []

        for ci, hdr in enumerate(headers):
            label = str(hdr) if hdr is not None else _col_letter(ci + 1)
            info = {
                "number_format":       None,
                "format_meaning":      "general",
                "fill_color_rgb":      None,
                "fill_meaning":        "none",
                "header_bold":         False,
                "hidden":              col_dims.get(ci+1, {}).get("hidden", False),
                "width":               col_dims.get(ci+1, {}).get("width"),
            }
            try:
                hc = ws.cell(1, ci + 1)
                info["header_bold"] = bool(hc.font and hc.font.bold)
            except Exception:
                pass
            if ci < len(row2):
                try:
                    cell = row2[ci]
                    fmt = cell.number_format or "General"
                    info["number_format"]  = fmt
                    info["format_meaning"] = _fmt_meaning(fmt)
                    fc = cell.fill.fgColor
                    if fc.type == "rgb":
                        rgb = str(fc.rgb).upper()
                        info["fill_color_rgb"] = rgb
                        info["fill_meaning"]   = _colour_meaning(rgb)
                    elif fc.type == "theme":
                        info["fill_color_rgb"] = f"theme:{fc.theme}"
                except Exception:
                    pass
            sheet_fmt[label] = info

        result[sname] = sheet_fmt

    wb.close()
    return result


# ─────────────────────────────────────────────────────────────────────────────
# COMBINED GRAPH
# ─────────────────────────────────────────────────────────────────────────────

def build_graph(path: str, original_path: str = None) -> dict:
    """Full graph = structure pass + format pass merged.

    If a .db file exists alongside the workbook, db_path and sheet_tables
    are added to the graph so StorageBackend and excel_run_code can
    use them immediately without a separate ingest call.
    """
    g = inspect_structure(path)
    g["original_path"]    = original_path or path
    g["ai_workbook_path"] = path

    fmt_data = inspect_formatting(path)
    if "error" not in fmt_data:
        for sname, col_fmts in fmt_data.items():
            if sname in g["sheets"]:
                g["sheets"][sname]["column_formats"] = col_fmts

    # ── Attach db metadata if a SQLite db exists alongside the workbook ───────
    import sqlite3 as _sq, json as _json
    p   = Path(path)
    db  = p.parent / (p.stem.replace("_ai_workbook", "") + ".db")
    if db.exists():
        g["db_path"] = str(db)
        try:
            conn = _sq.connect(str(db))
            rows = conn.execute(
                "SELECT sheet_name, table_name, row_count, "
                "columns_json, join_candidates, canonical_map "
                "FROM _sheet_registry"
            ).fetchall()
            conn.close()
            sheet_tables = {}
            for (sheet, table, row_count,
                 cols_json, jc_json, canon_json) in rows:
                sheet_tables[sheet] = {
                    "table":           table,
                    "rows":            row_count,
                    "columns":         _json.loads(cols_json)  if cols_json  else [],
                    "join_candidates": _json.loads(jc_json)    if jc_json    else [],
                    "canonical_cols":  _json.loads(canon_json) if canon_json else {},
                }
            g["sheet_tables"] = sheet_tables
        except Exception:
            g["sheet_tables"] = {}
    else:
        g["db_path"]      = None
        g["sheet_tables"] = {}

    return g


# ─────────────────────────────────────────────────────────────────────────────
# MARKDOWN RENDERER (for LLM context)
# ─────────────────────────────────────────────────────────────────────────────

def render_graph_md(g: dict) -> str:
    L = []; a = L.append
    a(f"# Workbook Graph")
    a(f"**AI Workbook:** `{g.get('ai_workbook_path','?')}`")
    if g.get("original_path") != g.get("ai_workbook_path"):
        a(f"**Original (read-only):** `{g.get('original_path','?')}`")
    a(f"**Sheets ({len(g['sheet_names'])}):** {', '.join(g['sheet_names'])}\n")

    # SQLite database section — shown when db exists
    if g.get("db_path"):
        a(f"## SQLite Database")
        a(f"- **Path:** `{g['db_path']}`")
        st = g.get("sheet_tables", {})
        if st:
            a("- **Tables:**")
            for sname, info in st.items():
                cols_preview = ", ".join(
                    f"`{c['name']}`" for c in info.get("columns", [])[:5]
                )
                if len(info.get("columns", [])) > 5:
                    cols_preview += f" ... (+{len(info['columns'])-5} more)"
                jc = info.get("join_candidates", [])
                jc_str = (f" | join keys: {', '.join(f'`{c}`' for c in jc)}"
                          if jc else "")
                a(f"  - `{sname}` → table `{info['table']}` "
                  f"({info['rows']:,} rows){jc_str}")
                a(f"    Columns: {cols_preview}")
            a("")
            a("**SQL JOIN example** (using join_candidates):")
            sheets = list(st.keys())
            if len(sheets) >= 2:
                s1, s2 = sheets[0], sheets[1]
                t1 = st[s1]["table"]
                t2 = st[s2]["table"]
                jc1 = st[s1].get("join_candidates", [])
                jk  = jc1[0] if jc1 else "shared_column"
                a(f"```sql")
                a(f"SELECT a.*, b.*")
                a(f"FROM {t1} a")
                a(f"LEFT JOIN {t2} b ON a.[{jk}] = b.[{jk}]")
                a(f"```")
        a("")

    if g["named_ranges"]:
        a("## Named Ranges")
        for n, d in g["named_ranges"].items():
            a(f"- **`{n}`** → {', '.join(d)}")
        a("")

    if g["cross_sheet_refs"]:
        a("## Cross-Sheet Dependencies")
        for src, targets in g["cross_sheet_refs"].items():
            a(f"- **{src}** → {', '.join(targets)}")
        a("")

    a("---")
    for sn, s in g["sheets"].items():
        a(f"## Sheet: `{sn}`")
        a(f"- **Size:** {s['max_row']:,} rows × {s['max_col']} columns")
        if s["headers"]:
            a(f"- **Headers:** {s['headers']}")

        fmts = s.get("column_formats", {})
        if s["column_types"] or fmts:
            a("- **Column details:**")
            for hdr in s["headers"]:
                label = str(hdr) if hdr is not None else "?"
                dtype = s["column_types"].get(label, "?")
                fi    = fmts.get(label, {})
                parts = [f"type={dtype}"]
                fm = fi.get("format_meaning","")
                nf = fi.get("number_format","")
                fl = fi.get("fill_meaning","none")
                hid = fi.get("hidden", False)
                if fm and fm != "general": parts.append(f"format={fm}")
                if nf and nf not in ("General",None): parts.append(f"`{nf}`")
                if fl != "none": parts.append(f"fill={fl}")
                if hid: parts.append("⚠️ HIDDEN")
                if fi.get("header_bold"): parts.append("bold-header")
                a(f"  - `{label}`: {' | '.join(parts)}")

        if s["sample_rows"]:
            a("- **Sample (rows 2–4):**")
            for row in s["sample_rows"][:3]:
                a(f"  - {row}")

        if s["merged_cells"]:
            a(f"- **Merged cells:** {', '.join(s['merged_cells'][:6])}")
        if s["tables"]:
            a("- **Tables:**")
            for t in s["tables"]:
                a(f"  - `{t['name']}` @ `{t['ref']}` — {t['columns']}")

        f = s["formulas"]
        if f["total"]:
            a(f"- **Formulas:** {f['total']} cells, "
              f"{len(f['patterns'])} unique patterns")
            for p in f["patterns"][:15]:
                cells = ", ".join(p["cells"])
                if p["count"] > len(p["cells"]):
                    cells += f" +{p['count']-len(p['cells'])} more"
                a(f"  - `{p['example_cell']}` (×{p['count']}): `{p['formula']}`")
                if p["cross_sheet_refs"]:
                    a(f"    → refs: **{', '.join(p['cross_sheet_refs'])}**")
        a("")

    if g.get("vba"):
        a("---\n## VBA Modules")
        for mod, code in g["vba"].items():
            a(f"### `{mod}`\n```vba\n{code.strip()}\n```\n")

    a("---\n## Rules for AI Agent")
    a("1. All writes → ai_workbook path only, never original")
    a("2. Use exact header names — never guess")
    a("3. `fill=input_hardcoded` (blue) → safe to modify")
    a("4. `fill=link_cross_sheet` (green) → formula-driven, modify formula not value")
    a("5. `fill=assumption` (yellow) → key assumptions, flag before changing")
    a("6. `format=percentage` → values are 0–1 decimals (0.75 = 75%)")
    a("7. `format=date` → may be stored as serial integer")
    a("8. ⚠️ HIDDEN columns → exclude from aggregations unless explicitly asked")
    a("9. Files >50K rows → always chunk, never load full dataset")
    return "\n".join(L)